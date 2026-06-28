"""
fill_excel.py  --  financial_data.json を読み込み、財務分析Excelを生成する。

使い方:
  python fill_excel.py <company_dir_name>
  例: python fill_excel.py dentsu

ルール（絶対変更禁止）:
  - Sheet1 = テンプレートの1枚目をそのまま使用。書式・レイアウト・ラベルを一切変えない。
    データセルにのみ値を書き込む（直近5期 → 列 E〜I / 行 5〜18）。
  - Sheet2以降 = 損益計算書 / 貸借対照表 / CF計算書 / セグメント / 同業比較
    を独自スタイルで追加する。

テンプレートのSheet1レイアウト（固定）:
  Row5  : E=年度1 F=年度2 G=年度3 H=年度4 I=年度5  (直近5期)
  Row6  : ROA
  Row7  : ROE
  Row8  : 営業利益率
  Row9  : 自己資本比率
  Row10 : 総資産回転率
  Row11 : インタレスト・カバレッジ
  Row12 : 株価推移
  Row13 : 純資産成長率
  Row14 : PBR
  Row15 : セグメント別売上高（ラベル行）
  Row16 : セグメントサブ行（企業名で上書き）
  Row17 : 減損額
  Row18 : 債券格付
"""

import sys
import json
import re
import shutil
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl が必要です。pip install openpyxl で導入してください。")
    sys.exit(1)

# ─────────────────────────────────────────────
# パス設定
# ─────────────────────────────────────────────
BASE_DIR = Path(__file__).parent.parent  # company-analyzer/
TEMPLATE = BASE_DIR / "templates" / "financial_sheet_template.xlsx"

# テンプレートSheet1のデータ列（5列固定）
DATA_COLS = ["E", "F", "G", "H", "I"]

# テンプレートSheet1の指標→行番号マッピング（変更禁止）
SHEET1_ROWS = {
    "roa":              6,
    "roe":              7,
    "operating_margin": 8,
    "equity_ratio":     9,
    "asset_turnover":   10,
    "icr":              11,
    "stock_price":      12,
    "bv_growth":        13,
    "pbr":              14,
    # row 15: セグメントラベル行（ラベルのみ）
    # row 16: セグメントサブ行（企業別セグメント名と直近期売上）
    "impairment":       17,
    "bond_rating":      18,
}


def load_data(company: str) -> dict:
    json_path = BASE_DIR / "outputs" / company / "financial_data.json"
    if not json_path.exists():
        raise FileNotFoundError(f"financial_data.json が見つかりません: {json_path}")
    with open(json_path, encoding="utf-8") as f:
        return json.load(f)


def fmt(v, digits=1):
    """数値をフォーマット。Noneは '-' に変換。"""
    if v is None:
        return "-"
    if isinstance(v, float):
        return round(v, digits)
    return v


def last5(values: list) -> list:
    """リストの直近5件を返す。5未満なら全件。"""
    if not values:
        return []
    return values[-5:] if len(values) >= 5 else values


# ─────────────────────────────────────────────
# Sheet1: テンプレートそのまま使用・データのみ書き込む
# ─────────────────────────────────────────────

def fill_sheet1_template(ws, data: dict):
    """
    テンプレートSheet1の書式・ラベルを一切変えずに、
    直近5期のデータだけを所定セルへ書き込む。
    """
    years = data.get("fiscal_years", [])
    m = data.get("metrics", {})

    recent_years = last5(years)
    pad = 5 - len(recent_years)  # データが5年未満の場合の左パディング数

    def extract_year_int(fy_str):
        """'2023年12月期' → 2023 (整数)。数字だけならそのまま返す。"""
        if isinstance(fy_str, int):
            return fy_str
        m = re.search(r'(\d{4})', str(fy_str))
        return int(m.group(1)) if m else fy_str

    # ── 年度ヘッダー (Row 5) — テンプレートと同じ整数形式 ──
    for i in range(5):
        cell = ws[f"{DATA_COLS[i]}5"]
        if i < pad:
            cell.value = None
        else:
            cell.value = extract_year_int(recent_years[i - pad])

    # ── 数値指標行 (Row 6〜14, 17) ──
    numeric_keys = [
        "roa", "roe", "operating_margin", "equity_ratio", "asset_turnover",
        "icr", "stock_price", "bv_growth", "pbr", "impairment",
    ]
    for key in numeric_keys:
        row = SHEET1_ROWS[key]
        vals = last5(m.get(key) or [])
        val_pad = 5 - len(vals)
        for i in range(5):
            cell = ws[f"{DATA_COLS[i]}{row}"]
            if i < val_pad:
                cell.value = "-"
            else:
                cell.value = fmt(vals[i - val_pad])

    # ── 債券格付 (Row 18、文字列) ──
    ratings = last5(m.get("bond_rating") or [])
    r_pad = 5 - len(ratings)
    for i in range(5):
        cell = ws[f"{DATA_COLS[i]}18"]
        if i < r_pad:
            cell.value = "-"
        else:
            v = ratings[i - r_pad]
            cell.value = v if v else "-"

    # ── ROIC行 (Row 19) — テンプレート外の追加行 ──
    roic_vals = last5(m.get("roic") or [])
    r_pad2 = 5 - len(roic_vals)
    ws["C19"] = "ROIC(%)"
    for i in range(5):
        cell = ws[f"{DATA_COLS[i]}19"]
        if i < r_pad2:
            cell.value = "-"
        else:
            cell.value = fmt(roic_vals[i - r_pad2])

    # ── WACC (J19 にコメントとして付記) ──
    wacc = data.get("metrics", {}).get("wacc")
    if wacc is not None:
        ws["J19"] = f"WACC想定 {wacc}%（仮置き）"

    # ── セグメント行 (Row 15〜16) ──
    # Row15: ラベル行（テンプレートのD15=国内, F15=海外）は残す。
    # Row16: セグメント名（D・F・H）と直近期売上（E・G・I）を企業別に上書き。
    segments = data.get("segments", [])
    if segments:
        seg_label_cols = ["D", "F", "H"]  # セグメント名を入れる列
        seg_data_cols  = ["E", "G", "I"]  # 直近期売上を入れる列
        for i, seg in enumerate(segments[:3]):
            name = seg.get("name", f"セグ{i+1}")
            rev_list = seg.get("revenue") or []
            latest_rev = next((v for v in reversed(rev_list) if v is not None), None)
            ws[f"{seg_label_cols[i]}16"] = name
            ws[f"{seg_data_cols[i]}16"] = fmt(latest_rev) if latest_rev is not None else "-"
        # 使われなかった列をクリア
        for i in range(len(segments), 3):
            ws[f"{seg_label_cols[i]}16"] = ""
            ws[f"{seg_data_cols[i]}16"] = ""


# ─────────────────────────────────────────────
# スタイル定数（Sheet2以降で使用）
# ─────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
LABEL_FILL  = PatternFill("solid", fgColor="D6E4F7")
LABEL_FONT  = Font(bold=True, size=9)
DATA_FONT   = Font(size=9)
CENTER      = Alignment(horizontal="center", vertical="center")
THIN_SIDE   = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def style_header(cell, text):
    cell.value = text
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def style_label(cell, text):
    cell.value = text
    cell.font = LABEL_FONT
    cell.fill = LABEL_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = THIN_BORDER


def style_data(cell, value):
    cell.value = fmt(value)
    cell.font = DATA_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def write_time_series_sheet(ws, title: str, years: list, rows: list):
    """汎用タイムシリーズシート（Sheet2以降で使用）"""
    ws.title = title
    ws.column_dimensions["A"].width = 32
    for i, yr in enumerate(years):
        col = get_column_letter(i + 2)
        ws.column_dimensions[col].width = 12
        style_header(ws.cell(1, i + 2), yr)
    style_header(ws.cell(1, 1), "指標")
    for r_idx, row in enumerate(rows, start=2):
        label = row["label"] + (f' {row.get("unit", "")}' if row.get("unit") else "")
        style_label(ws.cell(r_idx, 1), label)
        for c_idx, v in enumerate(row.get("values") or [None] * len(years), start=2):
            style_data(ws.cell(r_idx, c_idx), v)
    ws.freeze_panes = "B2"


# ─────────────────────────────────────────────
# Sheet2以降: 追加シート
# ─────────────────────────────────────────────

def write_pl_sheet(ws, data: dict):
    years = data.get("fiscal_years", [])
    pl = data.get("pl", {})
    rows = [
        {"label": "売上高",      "unit": "(億円)", "values": pl.get("revenue")},
        {"label": "営業利益",    "unit": "(億円)", "values": pl.get("operating_profit")},
        {"label": "当期純利益",  "unit": "(億円)", "values": pl.get("net_income")},
        {"label": "支払利息",    "unit": "(億円)", "values": pl.get("interest_expense")},
    ]
    write_time_series_sheet(ws, "損益計算書", years, rows)


def write_bs_sheet(ws, data: dict):
    years = data.get("fiscal_years", [])
    bs = data.get("bs", {})
    rows = [
        {"label": "資産合計",             "unit": "(億円)", "values": bs.get("total_assets")},
        {"label": "純資産合計（株主資本）", "unit": "(億円)", "values": bs.get("equity")},
        {"label": "純有利子負債",          "unit": "(億円)", "values": bs.get("net_debt")},
    ]
    write_time_series_sheet(ws, "貸借対照表", years, rows)


def write_cf_sheet(ws, data: dict):
    years = data.get("fiscal_years", [])
    cf = data.get("cf", {})
    rows = [
        {"label": "営業活動によるキャッシュ・フロー", "unit": "(億円)", "values": cf.get("operating_cf")},
        {"label": "投資活動によるキャッシュ・フロー", "unit": "(億円)", "values": cf.get("investing_cf")},
        {"label": "財務活動によるキャッシュ・フロー", "unit": "(億円)", "values": cf.get("financing_cf")},
        {"label": "フリー・キャッシュ・フロー",       "unit": "(億円)", "values": cf.get("fcf")},
    ]
    write_time_series_sheet(ws, "CF計算書", years, rows)


def write_segments_sheet(ws, data: dict):
    years = data.get("fiscal_years", [])
    segments = data.get("segments", [])
    ws.title = "セグメント"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    for i, yr in enumerate(years):
        col = get_column_letter(i + 3)
        ws.column_dimensions[col].width = 12
        style_header(ws.cell(1, i + 3), yr)
    style_header(ws.cell(1, 1), "セグメント")
    style_header(ws.cell(1, 2), "種別")
    row_idx = 2
    for seg in segments:
        name = seg.get("name", "不明")
        for kind, key in [("売上高(億)", "revenue"), ("利益(億)", "profit")]:
            style_label(ws.cell(row_idx, 1), name)
            style_label(ws.cell(row_idx, 2), kind)
            vals = seg.get(key) or [None] * len(years)
            for c_idx, v in enumerate(vals, start=3):
                style_data(ws.cell(row_idx, c_idx), v)
            row_idx += 1
    ws.freeze_panes = "C2"


def write_peers_sheet(ws, data: dict):
    peers = data.get("peers", [])
    ws.title = "同業比較"
    headers = [
        "企業名", "自己資本利益率ROE(%)", "投下資本利益率ROIC(%)",
        "株価純資産倍率PBR(倍)", "EV/EBITDA(倍)", "売上高営業利益率(%)",
    ]
    keys = ["name", "roe", "roic", "pbr", "ev_ebitda", "operating_margin"]
    for c, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(c)].width = 20
        style_header(ws.cell(1, c), h)
    for r, peer in enumerate(peers, 2):
        for c, key in enumerate(keys, 1):
            v = peer.get(key)
            if c == 1:
                style_label(ws.cell(r, c), v or "-")
            else:
                style_data(ws.cell(r, c), v)


# ─────────────────────────────────────────────
# メイン
# ─────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("使い方: python fill_excel.py <company>")
        sys.exit(1)

    company = sys.argv[1]
    data = load_data(company)
    company_jp = data.get("company", company)

    out_dir = BASE_DIR / "outputs" / company
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{company_jp}_財務分析.xlsx"

    if not TEMPLATE.exists():
        print(f"ERROR: テンプレートが見つかりません: {TEMPLATE}")
        sys.exit(1)

    # ── テンプレートをコピー（Sheet1の書式を完全継承）──
    shutil.copy(TEMPLATE, out_path)
    wb = openpyxl.load_workbook(out_path)

    # Sheet1: テンプレートのまま使う（削除・再生成禁止）
    ws1 = wb.active  # Sheet1
    fill_sheet1_template(ws1, data)

    # Sheet2・Sheet3（テンプレートの空シート）を削除
    for name in list(wb.sheetnames):
        if name != ws1.title:
            del wb[name]

    # 追加シート（Sheet2以降）
    write_pl_sheet(wb.create_sheet(), data)
    write_bs_sheet(wb.create_sheet(), data)
    write_cf_sheet(wb.create_sheet(), data)
    write_segments_sheet(wb.create_sheet(), data)
    write_peers_sheet(wb.create_sheet(), data)

    wb.save(out_path)
    print(f"[OK] Excel: {out_path}")


if __name__ == "__main__":
    main()
